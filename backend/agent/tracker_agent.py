"""
tracker_agent.py
----------------
The Tracker Agent stores and summarizes application records.

It prefers MongoDB when configured, and falls back to the local Excel
tracking sheet for backwards compatibility.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config
from agent.base_agent import BaseAgent
from agent.submission_agent import ApplicationResult

try:
    from pymongo import MongoClient
    from pymongo.errors import PyMongoError
except ImportError:  # pragma: no cover - fallback path
    MongoClient = None
    PyMongoError = Exception


COLUMNS = [
    "job_id",
    "platform",
    "title",
    "company",
    "location",
    "work_mode",
    "salary",
    "url",
    "date_posted",
    "confidence_score",
    "date_applied",
    "status",
    "notes",
]

STATUS_COLORS = {
    "Applied": "4FC3F7",
    "DryRun": "B0BEC5",
    "Failed": "EF9A9A",
    "Under Review": "FFF176",
    "Interview": "A5D6A7",
    "Rejected": "EF9A9A",
    "Offer": "81C784",
}

APPLIED_STATUSES = {
    "Applied",
    "DryRun",
    "Under Review",
    "Interview",
    "Offer",
}


class TrackerAgent(BaseAgent):
    """Manages application tracking in MongoDB or Excel."""

    def __init__(self, run_config=None):
        super().__init__("TrackerAgent", run_config=run_config)
        self.excel_path = config.EXCEL_FILE_PATH
        self.mongo_client = None
        self.mongo_collection = None
        self.use_mongodb = False

        if config.MONGODB_URI and MongoClient is not None:
            self._init_mongodb()
        else:
            self._ensure_workbook()
            self.log.info(f"Tracker backend: Excel ({self.excel_path})")

    def _init_mongodb(self) -> None:
        """Initialize the MongoDB collection and index."""
        try:
            self.mongo_client = MongoClient(config.MONGODB_URI, serverSelectionTimeoutMS=5000)
            database = self.mongo_client[config.MONGODB_DB]
            self.mongo_collection = database[config.MONGODB_COLLECTION]
            self.mongo_collection.create_index("job_id", unique=True)
            self.use_mongodb = True
            self.log.info(
                f"Tracker backend: MongoDB "
                f"(db={config.MONGODB_DB}, collection={config.MONGODB_COLLECTION})"
            )
        except PyMongoError as e:
            self.log.warning(f"MongoDB init failed, falling back to Excel: {e}")
            self.mongo_client = None
            self.mongo_collection = None
            self.use_mongodb = False
            self._ensure_workbook()
            self.log.info(f"Tracker backend: Excel ({self.excel_path})")

    def _ensure_workbook(self) -> None:
        """Create the Excel file with headers if it does not exist yet."""
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)

        if self.excel_path.exists():
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1565C0")
            cell.alignment = Alignment(horizontal="center")

        col_widths = {
            "A": 36,
            "B": 12,
            "C": 30,
            "D": 25,
            "E": 20,
            "F": 12,
            "G": 15,
            "H": 50,
            "I": 14,
            "J": 10,
            "K": 14,
            "L": 15,
            "M": 40,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(self.excel_path)
        self.log.info(f"Created new tracking workbook: {self.excel_path}")

    def _job_to_document(self, app: ApplicationResult) -> dict:
        """Convert an application result into the stored record shape."""
        job = app.job
        return {
            "job_id": job.job_id,
            "platform": job.platform,
            "title": job.title,
            "company": job.company,
            "location": job.location,
            "work_mode": job.work_mode,
            "salary": job.salary,
            "url": job.url,
            "date_posted": job.date_posted,
            "confidence_score": job.confidence_score,
            "date_applied": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "status": app.result.value,
            "notes": app.notes,
        }

    def get_applied_ids(self) -> set[str]:
        """Return the set of previously tracked job IDs."""
        if self.use_mongodb:
            ids = set(
                self.mongo_collection.distinct(
                    "job_id",
                    {"status": {"$in": list(APPLIED_STATUSES)}},
                )
            )
            self.log.info(f"Found {len(ids)} existing application records in MongoDB.")
            return ids

        wb = load_workbook(self.excel_path)
        ws = wb.active
        ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[COLUMNS.index("status")] in APPLIED_STATUSES:
                ids.add(str(row[0]))
        self.log.info(f"Found {len(ids)} existing application records in Excel.")
        return ids

    def record_results(self, results: list[ApplicationResult]) -> None:
        """Persist a batch of application results."""
        if self.use_mongodb:
            added = 0
            for app in results:
                record = self._job_to_document(app)
                result = self.mongo_collection.update_one(
                    {"job_id": record["job_id"]},
                    {"$setOnInsert": record},
                    upsert=True,
                )
                if result.upserted_id is not None:
                    added += 1
            self.log.info(
                f"Saved {added} new application records to "
                f"{config.MONGODB_DB}.{config.MONGODB_COLLECTION}"
            )
            return

        wb = load_workbook(self.excel_path)
        ws = wb.active
        existing_ids = {
            str(row[0])
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[0] and row[COLUMNS.index("status")] in APPLIED_STATUSES
        }

        added = 0
        for app in results:
            record = self._job_to_document(app)
            if record["job_id"] in existing_ids:
                self.log.debug(f"Record exists, skipping: {record['job_id']}")
                continue

            next_row = ws.max_row + 1
            for col_idx, key in enumerate(COLUMNS, start=1):
                ws.cell(row=next_row, column=col_idx, value=record[key])

            status_cell = ws.cell(row=next_row, column=COLUMNS.index("status") + 1)
            status_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=STATUS_COLORS.get(record["status"], "FFFFFF"),
            )

            added += 1
            existing_ids.add(record["job_id"])

        wb.save(self.excel_path)
        self.log.info(f"Saved {added} new application records to {self.excel_path}")

    def update_status(self, job_id: str, new_status: str, notes: str = "") -> bool:
        """Update the status of an existing application row by job_id."""
        if self.use_mongodb:
            update_doc = {"status": new_status}
            if notes:
                update_doc["notes"] = notes
            result = self.mongo_collection.update_one(
                {"job_id": job_id},
                {"$set": update_doc},
            )
            if result.matched_count:
                self.log.info(f"Updated status for {job_id} -> {new_status}")
                return True
            self.log.warning(f"job_id not found in MongoDB: {job_id}")
            return False

        wb = load_workbook(self.excel_path)
        ws = wb.active
        status_col = COLUMNS.index("status") + 1
        notes_col = COLUMNS.index("notes") + 1

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == job_id:
                row[status_col - 1].value = new_status
                row[status_col - 1].fill = PatternFill(
                    fill_type="solid",
                    fgColor=STATUS_COLORS.get(new_status, "FFFFFF"),
                )
                if notes:
                    row[notes_col - 1].value = notes
                wb.save(self.excel_path)
                self.log.info(f"Updated status for {job_id} -> {new_status}")
                return True

        self.log.warning(f"job_id not found in Excel: {job_id}")
        return False

    def get_summary(self) -> dict[str, int]:
        """Return a status -> count summary."""
        if self.use_mongodb:
            summary: dict[str, int] = {}
            pipeline = [
                {"$group": {"_id": "$status", "count": {"$sum": 1}}},
            ]
            for row in self.mongo_collection.aggregate(pipeline):
                if row["_id"]:
                    summary[str(row["_id"])] = int(row["count"])
            self.log.info(f"Application summary: {summary}")
            return summary

        wb = load_workbook(self.excel_path)
        ws = wb.active
        status_col = COLUMNS.index("status") + 1
        summary: dict[str, int] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            status = row[status_col - 1]
            if status:
                summary[status] = summary.get(status, 0) + 1

        self.log.info(f"Application summary: {summary}")
        return summary

    def get_stats(self) -> dict[str, int]:
        """Return a stable summary dict with common statuses pre-filled."""
        stats = {
            "total": 0,
            "Applied": 0,
            "DryRun": 0,
            "Failed": 0,
            "Under Review": 0,
            "Interview": 0,
            "Rejected": 0,
            "Offer": 0,
        }

        summary = self.get_summary()
        for status, count in summary.items():
            stats[status] = stats.get(status, 0) + count
            stats["total"] += count

        self.log.info(f"Application stats: {stats}")
        return stats

    def get_applied_today(self) -> int:
        """Return how many applications were recorded today."""
        today = datetime.now().strftime("%Y-%m-%d")

        if self.use_mongodb:
            count = self.mongo_collection.count_documents(
                {"date_applied": {"$regex": f"^{today}"}}
            )
            self.log.info(f"Applications recorded today: {count}")
            return count

        wb = load_workbook(self.excel_path)
        ws = wb.active
        applied_col = COLUMNS.index("date_applied")
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            value = row[applied_col]
            if value and str(value).startswith(today):
                count += 1

        self.log.info(f"Applications recorded today: {count}")
        return count

    def get_backend_info(self) -> dict[str, str | bool]:
        """Return the active tracking backend information."""
        if self.use_mongodb:
            return {
                "type": "mongodb",
                "connected": True,
                "db": config.MONGODB_DB,
                "collection": config.MONGODB_COLLECTION,
            }

        return {
            "type": "excel",
            "connected": self.excel_path.exists(),
            "path": str(self.excel_path),
        }

    def get_recent_records(self, limit: int = 10) -> list[dict]:
        """Return recent application records for frontend history views."""
        limit = max(1, min(limit, 50))

        if self.use_mongodb:
            cursor = (
                self.mongo_collection.find(
                    {},
                    {"_id": 0},
                )
                .sort("date_applied", -1)
                .limit(limit)
            )
            return [dict(row) for row in cursor]

        wb = load_workbook(self.excel_path)
        ws = wb.active
        records: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = {
                key: row[idx] if idx < len(row) else None
                for idx, key in enumerate(COLUMNS)
            }
            records.append(record)

        records.sort(key=lambda item: str(item.get("date_applied", "")), reverse=True)
        return records[:limit]

    def run(self, results: list[ApplicationResult]) -> None:
        """Record application results and log a summary."""
        self.record_results(results)
        summary = self.get_summary()
        self.log.info(
            "=== Application Pipeline Summary ===\n"
            + "\n".join(f"  {k}: {v}" for k, v in summary.items())
        )
